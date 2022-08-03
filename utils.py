from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import pandas as pd
import datetime
import json


# Save File to Excel [OLD]
def save_file_old(main_dir, filename, data):
    # Creating Data Frame with Data from Dictionary
    final_df = pd.concat({k: pd.DataFrame(v).transpose() for k, v in data.items()}, sort=True, axis=1)

    save_filename = main_dir + filename + '.xlsx'
    print(f'Working Directory: {main_dir}')
    check_path = Path(save_filename)

    # Checking if file already exists
    if check_path.exists():
        print(f'{filename}.xlsx already exists, creating new sheet')
        book = load_workbook(save_filename)
        writer = pd.ExcelWriter(save_filename, engine='openpyxl')
        writer.book = book
        # Putting data into the Excel Sheet
        final_df.to_excel(writer, sheet_name=datetime.today().strftime('%m-%d-%y'))
        writer.save()
        writer.close()
    else:
        print(f'{filename}.xlsx doesnt exist, creating new file')
        # Putting data into the Excel Sheet
        final_df.to_excel(save_filename, sheet_name=datetime.today().strftime('%m-%d-%y'))


def create_json(data, filename):
    # Creating a json file with complete call data:
    with open(filename, 'w') as file:
        json.dump(data, file)


def load_json(filename):
    # Loading JSON file with data
    with open(filename, 'r') as file:
        data = json.load(file)
    return data


def create_excel(main_dir, filename, data):
    # Creating Data Frame with Data from JSON

    # Flatten JSON structure into a list
    temp_list = []
    for k, v in data.items():
        date = k
        for item in v:
            raw = item.split("-")
            direction = raw[0].strip()
            time = raw[1].strip()
            number = raw[2].strip()
            temp_list.append([date, direction, time, number])

    df = pd.DataFrame(temp_list, columns=['Date', 'Direction', 'Time', 'Number'], index=None)
    save_filename = main_dir + filename + '.xlsx'
    print(f'Working Directory: {main_dir}')
    check_path = Path(save_filename)

    # Checking if file already exists
    if check_path.exists():
        print(f'{filename}.xlsx already exists, creating new sheet')
        book = load_workbook(save_filename)
        writer = pd.ExcelWriter(save_filename, engine='openpyxl')
        writer.book = book
        # Putting data into the Excel Sheet
        df.to_excel(writer, sheet_name=datetime.datetime.today().strftime('%m-%d-%y'), index=False)
        writer.save()
        writer.close()
    else:
        print(f'{filename}.xlsx doesnt exist, creating new file')
        # Putting data into the Excel Sheet
        df.to_excel(save_filename, sheet_name=datetime.datetime.today().strftime('%m-%d-%y'), index=False)


def parse_time(time_string: str):
    """
    125716 remove last 2 digits 1257 + 200 = 1457 (- 1200) = 2:57 pm
    074433 remove last 2 digits 0744 + 200 = 0944 = 9:44 am
    """

    time_string = int(time_string[:-2])
    time_string += 200

    return datetime.datetime.strptime(str(time_string), '%H%M').strftime('%I:%M %p')


def parse_month(month, abbreviated):
    return datetime.datetime.strptime(month, '%m').strftime('%b' if abbreviated else '%B')


def parse_date(date):
    return datetime.datetime.strptime(date, '%Y%m%d').strftime('%m/%d/%Y')


def parse_call_number(number):
    if number.startswith("%") or number.startswith("_"):
        return "NA"
    else:
        return number
